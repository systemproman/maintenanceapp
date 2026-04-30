from nicegui import ui, app
from components.menu import build_menu, hide_page_loader
from auth import has_permission
from services import db
from pathlib import Path
from datetime import datetime
import io

TABELAS_EXPORT = {
    'ativos': 'Ativos / Equipamentos',
    'ordens_servico': 'Ordens de Serviço',
    'os_atividades': 'Atividades de OS',
    'os_materiais': 'Materiais de OS',
    'funcionarios': 'Funcionários',
    'equipes': 'Equipes',
    'usuarios': 'Usuários',
    'audit_logs': 'Log de Ações',
}

TABELAS_IMPORT = {
    'ativos': 'Ativos / Equipamentos',
    'funcionarios': 'Funcionários',
    'equipes': 'Equipes',
}


def safe_notify(message: str, type_: str = 'info', multi_line: bool = False) -> None:
    try:
        ui.notify(str(message), type=type_, multi_line=multi_line)
    except Exception:
        pass


def _nome_tabela(tabela: str) -> str:
    return TABELAS_EXPORT.get(str(tabela or ''), str(tabela or '-'))


def _xlsx_bytes(rows: list[dict], sheet_name: str = 'DADOS') -> bytes:
    try:
        from openpyxl import Workbook
    except Exception as ex:
        raise RuntimeError('Biblioteca openpyxl não encontrada. Inclua openpyxl no requirements.txt.') from ex

    wb = Workbook()
    ws = wb.active
    ws.title = str(sheet_name or 'DADOS')[:31]
    campos = sorted({str(k) for row in rows for k in (row or {}).keys()}) if rows else []
    if campos:
        ws.append(campos)
        for row in rows:
            ws.append([row.get(c) for c in campos])
    else:
        ws.append(['SEM_DADOS'])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _read_xlsx_rows(raw: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except Exception as ex:
        raise RuntimeError('Biblioteca openpyxl não encontrada. Inclua openpyxl no requirements.txt.') from ex

    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers_raw = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(h).strip() if h is not None else '' for h in headers_raw]
    rows = []
    for values in rows_iter:
        item = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            val = values[idx] if idx < len(values) else None
            item[header] = '' if val is None else val
        if any(str(v).strip() for v in item.values()):
            rows.append(item)
    return rows


def gestao_dados_page():
    if not has_permission('/gestao-dados', 'abrir_tela'):
        safe_notify('Você não possui permissão para acessar Dados.', type_='negative')
        ui.navigate.to('/home')
        return

    pode_importar = has_permission('/gestao-dados', 'criar') or has_permission('/gestao-dados', 'editar')
    pode_exportar = has_permission('/gestao-dados', 'exportar')
    pode_reverter = has_permission('/gestao-dados', 'excluir') or has_permission('/gestao-dados', 'admin_modulo') or has_permission('/gestao-dados', 'gerenciar_permissoes')

    ui.add_head_html('''
    <style>
        html, body, #app { width: 100%; height: 100%; margin: 0; overflow: hidden !important; }
        .nicegui-content { height: 100dvh !important; overflow: hidden !important; }
        .q-page { min-height: 0 !important; height: 100dvh !important; overflow: hidden !important; }
        .dados-log-table .q-table__middle { max-height: 36vh; overflow: auto; }
        .dados-log-table .q-table th { font-weight: 800; color: #334155; }
    </style>
    ''')

    tabela_logs = None

    with ui.row().classes('fsl-app-shell w-full h-screen no-wrap bg-slate-100 overflow-hidden'):
        build_menu('/gestao-dados')
        with ui.column().classes('flex-1 h-full min-h-0 p-4 gap-4 overflow-hidden'):
            with ui.card().classes('w-full rounded-2xl shadow-sm border-0 bg-white p-4 shrink-0'):
                ui.label('DADOS').classes('text-xl font-bold text-slate-800')
                ui.label('Exportação em XLSX, importação em massa por XLSX e reversão de cargas importadas.').classes('text-xs text-slate-500')

            with ui.column().classes('w-full flex-1 min-h-0 gap-4 overflow-hidden'):
                with ui.row().classes('w-full gap-4 shrink-0 items-stretch'):
                    with ui.card().classes('flex-1 min-w-[360px] rounded-2xl shadow-sm border-0 bg-white p-4 gap-3'):
                        ui.label('EXPORTAR DADOS').classes('text-base font-bold text-slate-800')
                        tabela_exp = ui.select(TABELAS_EXPORT, label='MODELO / TABELA', value='ativos').props('outlined dense options-dense').classes('w-full')
                        status_export = ui.label('Baixa uma planilha XLSX sem limite de linhas.').classes('text-xs text-slate-500')

                        def baixar():
                            if not pode_exportar:
                                safe_notify('Sem permissão para exportar.', type_='negative')
                                return
                            try:
                                rows = db.listar_tabela_generica(tabela_exp.value, None)
                                nome = f"{tabela_exp.value}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                                out = Path('exports')
                                out.mkdir(exist_ok=True)
                                path = out / nome
                                path.write_bytes(_xlsx_bytes(rows, tabela_exp.value))
                                status_export.set_text(f'{len(rows)} linha(s) exportada(s) em XLSX.')
                                ui.download(str(path), filename=nome)
                            except Exception as ex:
                                safe_notify(str(ex), type_='negative', multi_line=True)

                        ui.button('EXPORTAR XLSX', icon='download', on_click=baixar).props('unelevated').classes('bg-amber-500 text-black font-bold')

                    with ui.card().classes('flex-1 min-w-[420px] rounded-2xl shadow-sm border-0 bg-white p-4 gap-3'):
                        ui.label('IMPORTAR DADOS').classes('text-base font-bold text-slate-800')
                        ui.label('Selecione o modelo, escolha a planilha XLSX exportada/editada e suba os dados. Cada importação gera um lote reversível.').classes('text-xs text-slate-500')
                        with ui.row().classes('w-full gap-3 items-end'):
                            tabela_imp = ui.select(TABELAS_IMPORT, label='MODELO PARA IMPORTAR', value='ativos').props('outlined dense options-dense').classes('flex-1 min-w-[220px]')
                            modo = ui.select({'upsert': 'Atualizar se ID existir / inserir se não existir', 'insert': 'Somente inserir'}, label='MODO', value='upsert').props('outlined dense options-dense').classes('flex-1 min-w-[260px]')
                        resultado = ui.label('').classes('text-xs text-slate-600')

                        def carregar_lotes():
                            if tabela_logs is None:
                                return
                            try:
                                rows = db.listar_cargas_dados(500)
                            except Exception:
                                rows = []
                            tabela_logs.rows = [
                                {
                                    'data': r.get('created_at') or '-',
                                    'usuario': r.get('usuario_nome') or r.get('usuario_id') or '-',
                                    'importado': _nome_tabela(r.get('tabela')),
                                    'arquivo': r.get('arquivo_nome') or '-',
                                    'modo': r.get('modo') or '-',
                                    'linhas': r.get('linhas') or 0,
                                    'status': 'REVERTIDO' if int(r.get('revertido') or 0) else 'ATIVO',
                                    'batch_id': r.get('id') or '',
                                }
                                for r in rows
                            ]
                            tabela_logs.update()

                        def process_upload(e):
                            if not pode_importar:
                                safe_notify('Sem permissão para importar.', type_='negative')
                                return
                            try:
                                raw = e.content.read()
                                nome_arquivo = getattr(e, 'name', None) or getattr(e, 'filename', None) or 'importacao.xlsx'
                                rows = _read_xlsx_rows(raw)
                                if not rows:
                                    safe_notify('A planilha não possui linhas para importar.', type_='warning')
                                    return
                                usuario_nome = app.storage.user.get('name') or app.storage.user.get('username') or 'ADMIN'
                                res = db.importar_tabela_generica(
                                    tabela_imp.value,
                                    rows,
                                    modo.value,
                                    usuario_id=app.storage.user.get('usuario_id'),
                                    usuario_nome=usuario_nome,
                                    arquivo_nome=nome_arquivo,
                                )
                                lote = res.get('batch_id') or '-'
                                resultado.set_text(f"Importação concluída: {res.get('linhas', 0)} linha(s). Lote: {lote}")
                                safe_notify('CARGA IMPORTADA COM SUCESSO.', type_='positive')
                                carregar_lotes()
                            except Exception as ex:
                                safe_notify(str(ex), type_='negative', multi_line=True)

                        ui.upload(label='SELECIONAR E IMPORTAR XLSX', auto_upload=True, multiple=False, on_upload=process_upload).props('accept=.xlsx color=amber').classes('w-full')

                with ui.card().classes('w-full flex-1 min-h-0 rounded-2xl shadow-sm border-0 bg-white p-4 gap-3 overflow-hidden'):
                    with ui.row().classes('w-full items-center justify-between shrink-0'):
                        ui.label('IMPORTAÇÕES REALIZADAS').classes('text-base font-bold text-slate-800')
                        ui.button('ATUALIZAR', icon='refresh', on_click=lambda: carregar_lotes()).props('flat dense')

                    def reverter(batch_id: str):
                        if not pode_reverter:
                            safe_notify('Sem permissão para reverter carga.', type_='negative')
                            return
                        try:
                            res = db.reverter_carga_dados(batch_id, usuario_id=app.storage.user.get('usuario_id'))
                            safe_notify(f"Carga revertida: {res.get('linhas', 0)} registro(s).", type_='positive')
                            carregar_lotes()
                        except Exception as ex:
                            safe_notify(str(ex), type_='negative', multi_line=True)

                    columns = [
                        {'name': 'data', 'label': 'DATA', 'field': 'data', 'align': 'left'},
                        {'name': 'usuario', 'label': 'USUÁRIO', 'field': 'usuario', 'align': 'left'},
                        {'name': 'importado', 'label': 'O QUE FOI IMPORTADO', 'field': 'importado', 'align': 'left'},
                        {'name': 'arquivo', 'label': 'ARQUIVO', 'field': 'arquivo', 'align': 'left'},
                        {'name': 'modo', 'label': 'MODO', 'field': 'modo', 'align': 'left'},
                        {'name': 'linhas', 'label': 'LINHAS', 'field': 'linhas', 'align': 'right'},
                        {'name': 'status', 'label': 'STATUS', 'field': 'status', 'align': 'left'},
                        {'name': 'acao', 'label': 'REVERSÃO', 'field': 'batch_id', 'align': 'center'},
                    ]
                    tabela_logs = ui.table(columns=columns, rows=[], row_key='batch_id', pagination=12).classes('w-full dados-log-table')
                    tabela_logs.add_slot('body-cell-acao', '''
                        <q-td :props="props">
                          <q-btn v-if="props.row.status !== 'REVERTIDO'" dense unelevated color="negative" icon="undo" label="REVERTER" @click="$parent.$emit('reverter', props.row.batch_id)" />
                          <q-badge v-else color="grey" label="REVERTIDO" />
                        </q-td>
                    ''')
                    tabela_logs.on('reverter', lambda e: reverter(e.args))
                    carregar_lotes()

    hide_page_loader()
